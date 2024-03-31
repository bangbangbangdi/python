# -------------------- excel操作 --------------------
# 对excel文件进行读写操作需要用到 openpyxl 的第三方库
import datetime
import random

import openpyxl
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference, BarChart3D, PieChart
from copy import deepcopy


# -------------------- 读操作 --------------------
def read_excel():
    # 加载一个工作簿(整个Excel文件)
    wb = openpyxl.load_workbook('./testFile/2020年销售数据.xlsx')

    # 获取所有工作表的名字(所有的sheet)
    print(wb.sheetnames)

    # 获取第一个sheet
    sheet = wb.worksheets[0]
    # 获取单元格的范围
    print(sheet.dimensions)
    # 获取单元格的最大行、最大列
    print(sheet.max_row, sheet.max_column)

    # 通过cell获取指定单元格的数据
    print(sheet.cell(3, 3).value)
    # 通过指定单元格坐标获取对应单元格
    print(sheet['D1'].value)

    print('-------获取多个单元格----------')
    # 获取多个单元格
    print(sheet['A1:G2'])

    print('-------读取多个单元格的数据----------')
    # 读取多个单元格的数据
    for cells in sheet['A1:G2']:
        for cell in cells:
            print(cell.value, end=',')
        print()

    print('-------读取所有单元格的数据----------')
    # 按行遍历所有数据
    # for i in range(1, sheet.max_row + 1):
    for i in range(1, 10):
        for j in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=i, column=j)
            # print(cell.value, end='\t')

            # 格式化数据
            # print(cell.data_type, end=';')
            # print(type(cell), end=';')
            if j == 1 and cell.data_type != 'd':
                print(f'{cell.value:<10}', end='\t')
            elif j == 1:
                fm_val = cell.value.strftime('%Y/%m/%d')
                print(f'{fm_val:<10}', end='\t')
            elif j == 3:
                print(f'{cell.value:<6}', end='\t')
            elif j == 6:
                print(f'{cell.value:<6}', end='\t')
            else:
                print(cell.value, end='\t')
        print()


# -------------------- 写操作 --------------------
def write_excel():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'characters'
    titles = ('姓名', '力量', '敏捷', '耐力', '幸运', '成长', '财富')

    # 我们很多时候会需要获取遍历对象的索引值
    col_index = 0
    for title in titles:
        col_index += 1
        sheet.cell(row=1, column=col_index, value=title)
    # 不过像上面这样每次都要单独创建一个变量并手动的自增好麻烦...
    # 有没有更优雅的写法呢?

    # for col_index, title in enumerate(titles):
    #     sheet.cell(row=1, column=col_index + 1, value=title)

    names = ('kino', 'kirii', 'erms', 'shibo', 'kuroniko', 'kaiba')
    for row_index, name in enumerate(names):
        row = row_index + 2
        for col in range(1, 1 + len(titles)):
            # 名字需要单独设置
            if col == 1:
                sheet.cell(row=row, column=col, value=name)
            else:
                sheet.cell(row=row, column=col, value=random.randint(1, 101))
            # 这个if-else看着好不顺眼...明明两个操作都是要给对应的单元格赋值
            # 就不能简单一点吗?

            # sheet.cell(row=row, column=col, value=name if col == 1 else random.randint(1, 101))

    wb.save('./testFile/CharacterStatus.xlsx')


# -------------------- 使用excel公式 --------------------
def use_excel_func():
    """
    在CharacterStatus.xlsx的最右边添加一列 平均值 计算角色的平均属性
    :return:
    """
    wb = openpyxl.load_workbook('./testFile/CharacterStatus.xlsx')
    sheet = wb.worksheets[0]
    max_row = sheet.max_row
    max_col = sheet.max_column
    # 这里我们需要把列数转为对应的大写字母  例如 1 -> A; 2 -> B
    avg_start = 'B'
    avg_end = chr(max_col + 64)
    avg = chr(max_col + 64 + 1)
    for row in range(1, max_row + 1):
        sheet[f'{avg}{row}'] = f'=average({avg_start}{row}:{avg_end}{row})' if row != 1 else '平均值'
    wb.save('./testFile/CharacterStatus.xlsx')


# -------------------- excel格式化 --------------------
# 需求:
# 1.将平均数从小数转为整数
# 2.为整个表格添加边框
# 3.表头加粗
# 4.平均数列添加背景颜色
def format_excel():
    file_path = './testFile/CharacterStatus.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.worksheets[0]
    max_row = sheet.max_row
    max_col = sheet.max_column

    # 边框线条 请告诉我颜色(16进制RGB)以及线条样式
    side = Side(color='000000', style='thin')
    # 设置边框
    border = Border(left=side, right=side, top=side, bottom=side)
    # 设置字体
    font = Font(bold=True)
    # 设置背景颜色 patternType = 填充模式(solid指纯色填充) , fgColor = 前景色(花纹颜色) , bgColor 背景色
    pattern = PatternFill(patternType='solid', fgColor='00F5FF')
    # pattern = PatternFill(patternType='darkDown', fgColor='00F5FF', bgColor='ff4500')

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            sheet.cell(row=row, column=col).border = border
            if row == 1:
                # 表头加粗
                sheet.cell(row=row, column=col).font = font
            if col == max_col:
                # 将平均数的列设置为整数: 这里的'0'实际上是格式化代码 -> 更多可以Google或者直接查看文档gg
                sheet.cell(row=row, column=max_col).number_format = '0'
                sheet.cell(row=row, column=max_col).fill = pattern

    wb.save(file_path)


# -------------------- 绘制表格 -------------------
def create_chart():
    file_path = './testFile/CharacterStatus.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.worksheets[0]
    max_row = sheet.max_row
    max_col = sheet.max_column

    # -------------------生成垂直柱状图
    # 获取BarChart(图表)对象
    chart1 = BarChart()
    # 设置图表类型 col垂直柱状图
    chart1.type = 'col'
    # 图表类型
    chart1.style = 10
    # 设置图表大小
    chart1.height = 9
    chart1.width = 18
    # 设置表头
    chart1.title = 'CharacterStatus'
    # 设置x轴标题
    chart1.x_axis.title = 'Status'
    # 设置y轴标题
    chart1.y_axis.title = 'Value'

    # 设置数据
    data = Reference(sheet, min_col=2, min_row=1, max_col=max_col, max_row=max_row)
    # 设置类别 category
    cats = Reference(sheet, min_col=1, min_row=2, max_row=max_row)
    # 将数据添加到图表中,titles_from_date表示数据源是否有标题(第一行)
    chart1.add_data(data, titles_from_data=True)
    # 将类别添加到图表对象中
    chart1.set_categories(cats)
    # 将设置好的图表插入到指定位置
    sheet.add_chart(chart1, 'A10')

    # -----------------生成水平柱状图
    # 如果我们想要生成一个新的图表的话,无疑需要再创建一个图表对象,然后再进行一系列跟上述相同的的设置
    # 好麻烦...能不能从A对象copy出一个相同的B对象呢? 这样我们就能省去很多设置的操作了
    # 当然可以~
    chart2 = deepcopy(chart1)
    chart2.style = 11
    chart2.type = 'bar'
    chart2.title = 'Horizontal chart'
    sheet.add_chart(chart2, 'N10')

    # ----------------- 生成饼状图
    chart3 = PieChart()
    chart3.title = 'Kino PieChart'
    labels = Reference(sheet, min_col=2, min_row=1, max_col=max_col)
    pie_data = Reference(sheet, min_col=1, min_row=2, max_col=max_col)
    # from_rows 从行当中获取数据
    chart3.add_data(pie_data, titles_from_data=True, from_rows=True)
    chart3.set_categories(labels)
    sheet.add_chart(chart3, 'A27')

    # slice = DataPoint(idx=2, explosion=20)
    # chart3.series[0].data_points = [slice]

    # ----------------- 生成3D柱状图
    chart4 = BarChart3D()
    chart4.title = '3D BarChart'
    chart4.height = 9
    chart4.width = 18
    chart4.add_data(data, titles_from_data=True)
    chart4.set_categories(cats)
    sheet.add_chart(chart4, 'N27')

    wb.save(file_path)


# -------------------- main --------------------

def main():
    # read_excel()
    # write_excel()
    # use_excel_func()
    # format_excel()
    create_chart()


if __name__ == '__main__':
    main()
